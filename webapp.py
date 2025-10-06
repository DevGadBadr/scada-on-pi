from flask import Flask,render_template,send_file,request
from flask_socketio import SocketIO, emit
import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
import json 
from datetime import datetime,timezone
from zoneinfo import ZoneInfo
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.dates as mdates
import os
from collections import defaultdict
import re 
from pathlib import Path
from urllib.parse import quote

load_dotenv()

ISTANBUL = ZoneInfo("Europe/Istanbul")
app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

connection = psycopg2.connect(
    host="localhost",
    port="5432",
    database="tpmdb",
    password="scadaonpi",
    user="devgadbadr"
)
connection.autocommit = True
cursor = connection.cursor()

def calculate_gen_hours(rows, end_at=None):
    """
    rows: iterable of (status, timestamp)
    returns: {"Generator 1": 5.0, "Generator 2": 3.0, ...}  (hours)
    """
    GEN_RE = re.compile(r"gen\s*(\d+)\s+(on|off)", re.IGNORECASE)
    last_on = {}
    totals = defaultdict(float)
    if end_at is None:
        end_at = datetime.now(timezone.utc)

    for status, ts in sorted(rows, key=lambda r: r[1]):
        m = GEN_RE.search(str(status))
        if not m:
            continue
        gen_id = int(m.group(1))
        state = m.group(2).lower()

        if state == "on":
            last_on[gen_id] = ts
        elif state == "off" and gen_id in last_on:
            delta = (ts - last_on[gen_id]).total_seconds()
            if delta > 0:
                totals[gen_id] += delta
            last_on.pop(gen_id, None)

    # close any open ON periods
    for gen_id, on_ts in last_on.items():
        delta = (end_at - on_ts).total_seconds()
        if delta > 0:
            totals[gen_id] += delta

    return {f"Generator {gid}": round(sec / 3600, 3) for gid, sec in totals.items()}

def _parse_iso_aware(s: str) -> datetime:
    # accepts "...Z" or "+00:00" or naive (assume UTC)
    if isinstance(s, datetime):
        dt = s if s.tzinfo else s.replace(tzinfo=timezone.utc)
    else:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
    return dt

def _to_tr_naive(dt: datetime) -> datetime:
    # to Europe/Istanbul, then strip tzinfo for matplotlib/xlsx friendliness
    return dt.astimezone(ISTANBUL).replace(tzinfo=None)

def _numify(v):
    if isinstance(v, (int, float)) or v is None:
        return v
    if isinstance(v, str):
        try:
            return int(v) if v.strip().isdigit() else float(v)
        except Exception:
            return v
    return v

def make_charts_pdf(
    timestamps,                         # list[str|datetime]
    total_active_power,                 # list[number]
    output_path,                        # "/home/.../report.pdf"
    from_iso=None,
    to_iso=None,
    last_row=None,                      # {"data": {...}} or {...}
    genhours=None,                      # {'Generator 1': 5.0, ...}
    last_ts=None                        # ISO string or datetime for "Last readings at ..."
):
    if not (timestamps and total_active_power):
        raise ValueError("No data provided")
    if len(timestamps) != len(total_active_power):
        raise ValueError("timestamps and power must have equal length")

    Path(os.path.dirname(output_path) or ".").mkdir(parents=True, exist_ok=True)

    ts_tr = [_to_tr_naive(_parse_iso_aware(t)) for t in timestamps]

    # subtitle "from .. to .. (Europe/Istanbul)" if provided
    if from_iso and to_iso:
        f_tr = _to_tr_naive(_parse_iso_aware(from_iso)).strftime("%Y-%m-%d %H:%M")
        t_tr = _to_tr_naive(_parse_iso_aware(to_iso  )).strftime("%Y-%m-%d %H:%M")
        subtitle = f"from {f_tr} to {t_tr} (Europe/Istanbul)"
    else:
        subtitle = "Europe/Istanbul"

    with PdfPages(output_path) as pdf:
        # ---- Page 1: Working hours bar chart ----
        order = ["Generator 1", "Generator 2", "Generator 3"]
        gh = genhours or {}
        values = [_numify(gh.get(name, 0)) for name in order]
        max_v = max(values) if values else 0

        fig0, ax0 = plt.subplots(figsize=(11.69, 8.27))
        fig0.suptitle("TPM-04ES Report", fontsize=18, fontweight="bold", y=0.98)
        ax0.set_title("Generators Working Hours\n" + subtitle, fontsize=13, pad=12)

        ax0.barh(order, values, height=0.45)
        ax0.set_xlabel("Hours")
        ax0.set_ylabel("Generator")
        ax0.grid(True, axis="x", linestyle="--", alpha=0.4)
        ax0.margins(x=0.10, y=0.20)
        if max_v > 0:
            ax0.set_xlim(0, max_v * 1.15)
        for y, v in enumerate(values):
            ax0.text(v, y, f"  {v:g}", va="center", ha="left")

        fig0.subplots_adjust(left=0.12, right=0.96, top=0.90, bottom=0.12)
        plt.tight_layout(pad=1.2, rect=[0, 0, 1, 0.95])
        pdf.savefig(fig0); plt.close(fig0)

        # --- Page 2: Total Active Power ---
        fig1, ax1 = plt.subplots(figsize=(11.69, 8.27))
        ax1.plot(ts_tr, total_active_power)
        ax1.set_title("Total Active Power\n" + subtitle, fontsize=14, pad=10)
        ax1.set_xlabel("Time (Europe/Istanbul)")
        ax1.set_ylabel("Total Active Power (W)")
        ax1.grid(True, linestyle="--", alpha=0.4)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d\n%H:%M"))
        fig1.autofmt_xdate()
        ax1.margins(x=0.03, y=0.10)
        fig1.subplots_adjust(left=0.10, right=0.97, top=0.90, bottom=0.18)
        plt.tight_layout(pad=1.2)
        pdf.savefig(fig1); plt.close(fig1)

        # --- Page 3: Last readings table ---
        data_dict = {}
        if isinstance(last_row, dict):
            if "data" in last_row and isinstance(last_row["data"], dict):
                data_dict = last_row["data"]
            else:
                data_dict = last_row  # readings dict directly

        keys = sorted(data_dict.keys()) if data_dict else []
        table_rows = [[k, _numify(data_dict.get(k))] for k in keys]

        # use passed last_ts for title if provided
        if last_ts:
            try:
                title_ts = _to_tr_naive(_parse_iso_aware(last_ts)).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                title_ts = str(last_ts)
        else:
            title_ts = "unknown"

        fig3, ax3 = plt.subplots(figsize=(11.69, 8.27))
        ax3.axis("off")
        ax3.set_title(f"Last readings at {title_ts} (Europe/Istanbul)", fontsize=14, pad=16)

        if table_rows:
            tbl = ax3.table(
                cellText=table_rows,
                colLabels=["Reading", "Value"],
                cellLoc="center",
                colWidths=[0.55, 0.30],
                loc="upper center",
            )
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(9)
            tbl.scale(1.05, 1.12)
        else:
            ax3.text(0.5, 0.5, "No readings available", ha="center", va="center", fontsize=12)

        fig3.subplots_adjust(left=0.06, right=0.94, top=0.92, bottom=0.08)
        plt.tight_layout(pad=1.0)
        pdf.savefig(fig3); plt.close(fig3)

    return output_path


def parse_iso_to_utc(iso_str: str) -> datetime:
    # "Z" means UTC; make it ISO8601-friendly for fromisoformat
    if iso_str.endswith("Z"):
        iso_str = iso_str[:-1] + "+00:00"
    dt = datetime.fromisoformat(iso_str)
    # ensure tz-aware UTC
    return dt.astimezone(timezone.utc)

def to_excel_naive(dt: datetime, zone: ZoneInfo = ISTANBUL) -> datetime:
    if dt is None:
        return None
    if dt.tzinfo is None:
        # already naive → keep as-is (or assume UTC, up to you)
        return dt
    # convert to your display zone, then strip tzinfo
    return dt.astimezone(zone).replace(tzinfo=None)

@app.route("/")
def main():
    return render_template("index.html")

@app.route("/downloadlog",methods=["POST"])
def donwload_log():
    print("body is:",request.get_json())
    datafilter = request.get_json()
    # cursor.execute("select * from tpmreading order by timestamp desc")
    fromm = parse_iso_to_utc(datafilter['from'])
    to = parse_iso_to_utc(datafilter['to'])
    cursor.execute("""
        SELECT *
        FROM tpmreading
        WHERE "timestamp" BETWEEN %s AND %s
        ORDER BY "timestamp" DESC
    """, (fromm, to))
    data = cursor.fetchall()

    cursor.execute("""
        SELECT status, "timestamp"
        FROM gens
        WHERE "timestamp" BETWEEN %s AND %s
        ORDER BY "timestamp" DESC
    """, (fromm, to))
    gensdata = cursor.fetchall()

    genhours = calculate_gen_hours(gensdata)
    print(genhours)

    readings_rows = [
        {"timestamp": row[2], "data": dict(row[1])} for row in data
    ]

    # Collect all JSON keys across all rows
    all_keys = []
    seen = set()
    for r in data:
        payload = r[1]
        if isinstance(payload, str):
            payload = json.loads(payload)
        if isinstance(payload, dict):
            for k in payload.keys():
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TPM Report"
    header = ["timestamp"] + all_keys
    ws.append(header)

     # Write data rows
    for r in data:
        payload = r[1]
        if isinstance(payload, str):
            payload = json.loads(payload)
        payload = payload or {}

        row = [to_excel_naive(r[2])]
        for k in all_keys:
            v = payload.get(k)
            # try to coerce numeric strings to numbers
            if isinstance(v, str):
                try:
                    if v.strip().isdigit():
                        v = int(v)
                    else:
                        v = float(v)
                except Exception:
                    pass
            row.append(v)
        ws.append(row)

    # (nice) auto width
    for col_idx, col_name in enumerate(header, start=1):
        max_len = len(str(col_name))
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_idx).value
            max_len = max(max_len, len(str(cell_val)) if cell_val is not None else 0)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    timestamps = [x[2] for x in data]
    activePowers = []
    activeEnergies = []
    for row in data:
        dataDict = row[1]
        activePower = dataDict["Total Active Power"]
        activeEnergy =  dataDict["Total Active Import Energy"]
        activePowers.append(activePower)
        activeEnergies.append(activeEnergy)

    if datafilter['file'] == 'excel':
        xlsx_path = f"/home/bigled/scadaonpi/reports/tpm_report_{now}.xlsx"
        wb.save(xlsx_path)
        print(f"✅ Exported {len(data)} rows → {xlsx_path}")
        return send_file(xlsx_path, as_attachment=True, download_name=f"tpm_report_{now}.xlsx", max_age=0)
    elif datafilter['file'] == 'pdf':
        pdfOutput = f"/home/bigled/scadaonpi/reports/tpm_report_{now}.pdf"
        make_charts_pdf(timestamps,activePowers,pdfOutput,fromm,to,readings_rows[0]['data'],genhours,readings_rows[0]['timestamp'])
        print(f"✅ PDF saved: {pdfOutput}")
        filename = Path(pdfOutput).name
        rv = send_file(
            pdfOutput,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf",
            conditional=False,          # <- important: avoids 304/Range → 0 B
            max_age=0
        )
        # Strongly disable caches and help IDM/browser naming
        rv.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        rv.headers["Pragma"] = "no-cache"
        rv.headers["Expires"] = "0"
        # Supply both filename and filename* (UTF-8) explicitly
        rv.headers["Content-Disposition"] = (
            f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"
        )
        return rv

@socketio.on("modbus-data")
def dataReceived(payload):
    emit("modbus-data",payload,broadcast=True)

if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=3000,allow_unsafe_werkzeug=True)
